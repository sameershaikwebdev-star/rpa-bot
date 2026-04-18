"""
Report Generator
Takes the bot's result list and produces:
  1. A styled HTML report (reports/report_<timestamp>.html)
  2. An Excel summary (reports/report_<timestamp>.xlsx)

FIXES:
  1. _build_dataframe() — Status/Timestamp/Error are set AFTER row_data.update()
                          so row_data keys can never overwrite them
  2. _build_dataframe() — internal keys (_validation_errors, _row_index) are
                          stripped from row_data before they reach any report column
"""

import pandas as pd
import logging
from datetime import datetime
from pathlib import Path

logger = logging.getLogger(__name__)

REPORT_DIR = Path("reports")
REPORT_DIR.mkdir(exist_ok=True)

# Keys added internally by main.py validation — must never appear as report columns
_INTERNAL_KEYS = {"_validation_errors", "_row_index"}


class ReportGenerator:
    """Generates HTML + Excel reports from RPA bot results."""

    def __init__(self, results: list):
        self.results = results
        self.df      = self._build_dataframe()
        self.ts      = datetime.now().strftime("%Y%m%d_%H%M%S")

    # ─── Build DataFrame ──────────────────────────────────────────────────────

    def _build_dataframe(self) -> pd.DataFrame:
        rows = []
        for r in self.results:
            # FIX 2: strip internal tracking keys from row_data before building columns
            row_data = {
                k: v for k, v in r.get("row_data", {}).items()
                if k not in _INTERNAL_KEYS
            }

            # Start with row_data fields
            row = dict(row_data)

            # FIX 1: set Status/Timestamp/Error AFTER spreading row_data so they
            # can never be overwritten by a coincidentally named row_data column
            row["Status"]    = r.get("status", "UNKNOWN")
            row["Timestamp"] = r.get("timestamp", "")
            row["Error"]     = r.get("error", "") or ""

            rows.append(row)

        df = pd.DataFrame(rows)

        # Move Status, Timestamp, Error to the front for readability
        priority_cols = ["Status", "Timestamp", "Error"]
        other_cols    = [c for c in df.columns if c not in priority_cols]
        return df[priority_cols + other_cols]

    # ─── Summary Stats ────────────────────────────────────────────────────────

    def _stats(self) -> dict:
        total   = len(self.df)
        success = (self.df["Status"] == "SUCCESS").sum()
        failed  = total - success
        rate    = round(success / total * 100, 1) if total else 0
        return {
            "total":   total,
            "success": int(success),
            "failed":  int(failed),
            "rate":    rate,
        }

    # ─── Excel Report ─────────────────────────────────────────────────────────

    def to_excel(self) -> str:
        path  = REPORT_DIR / f"report_{self.ts}.xlsx"
        stats = self._stats()

        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            self.df.to_excel(writer, sheet_name="Results", index=False)

            summary_df = pd.DataFrame([stats])
            summary_df.columns = ["Total", "Success", "Failed", "Success Rate (%)"]
            summary_df.to_excel(writer, sheet_name="Summary", index=False)

            # Auto-size columns in Results sheet
            ws = writer.sheets["Results"]
            for col in ws.columns:
                max_len = max(len(str(cell.value or "")) for cell in col) + 2
                ws.column_dimensions[col[0].column_letter].width = min(max_len, 40)

            # Colour-code Status column: green for SUCCESS, red for FAILED/ERROR
            from openpyxl.styles import PatternFill, Font
            green_fill = PatternFill("solid", fgColor="22c55e")
            red_fill   = PatternFill("solid", fgColor="ef4444")
            white_font = Font(color="FFFFFF", bold=True)

            status_col_idx = self.df.columns.get_loc("Status") + 1  # 1-based
            for row_cells in ws.iter_rows(min_row=2, min_col=status_col_idx,
                                          max_col=status_col_idx):
                for cell in row_cells:
                    if str(cell.value) == "SUCCESS":
                        cell.fill = green_fill
                        cell.font = white_font
                    elif cell.value in ("FAILED", "ERROR"):
                        cell.fill = red_fill
                        cell.font = white_font

        logger.info(f"📊 Excel report saved: {path}")
        return str(path)

    # ─── HTML Report ──────────────────────────────────────────────────────────

    def to_html(self) -> str:
        path  = REPORT_DIR / f"report_{self.ts}.html"
        stats = self._stats()

        rows_html = ""
        for _, row in self.df.iterrows():
            status = row.get("Status", "")
            color  = "#22c55e" if status == "SUCCESS" else "#ef4444"
            badge  = (
                f'<span style="background:{color};color:#fff;padding:2px 10px;'
                f'border-radius:999px;font-size:11px;font-weight:700">{status}</span>'
            )
            cols = "".join(
                f"<td style='padding:8px 14px;border-bottom:1px solid #1e293b'>"
                f"{badge if c == 'Status' else str(row[c])}</td>"
                for c in self.df.columns
            )
            rows_html += f"<tr>{cols}</tr>\n"

        headers = "".join(
            f"<th style='padding:10px 14px;text-align:left;"
            f"background:#0f172a;color:#94a3b8;font-size:11px;"
            f"text-transform:uppercase;letter-spacing:1px'>{c}</th>"
            for c in self.df.columns
        )

        html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<title>RPA Bot Report</title>
<link href="https://fonts.googleapis.com/css2?family=Space+Mono:wght@400;700&family=Inter:wght@400;500;600&display=swap" rel="stylesheet">
<style>
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{ background: #020817; color: #e2e8f0; font-family: 'Inter', sans-serif; padding: 40px; }}
  h1 {{ font-family: 'Space Mono', monospace; font-size: 28px; color: #38bdf8;
        border-left: 4px solid #38bdf8; padding-left: 16px; margin-bottom: 8px; }}
  .subtitle {{ color: #475569; font-size: 13px; margin-bottom: 36px; }}
  .cards {{ display: flex; gap: 16px; margin-bottom: 40px; flex-wrap: wrap; }}
  .card {{ flex: 1; min-width: 140px; background: #0f172a; border: 1px solid #1e293b;
           border-radius: 12px; padding: 20px 24px; }}
  .card-label {{ font-size: 11px; color: #64748b; text-transform: uppercase;
                 letter-spacing: 1px; margin-bottom: 8px; }}
  .card-value {{ font-family: 'Space Mono', monospace; font-size: 36px; font-weight: 700; }}
  .green {{ color: #22c55e; }} .red {{ color: #ef4444; }}
  .blue  {{ color: #38bdf8; }} .yellow{{ color: #facc15; }}
  .table-wrap {{ overflow-x: auto; border-radius: 12px; border: 1px solid #1e293b; }}
  table  {{ width: 100%; border-collapse: collapse; background: #0f172a; white-space: nowrap; }}
  tr:hover td {{ background: #1e293b44; }}
  .footer {{ margin-top: 32px; color: #334155; font-size: 12px; font-family: 'Space Mono', monospace; }}
</style>
</head>
<body>
<h1>⚡ RPA BOT — Execution Report</h1>
<p class="subtitle">Generated: {datetime.now().strftime('%B %d, %Y at %H:%M:%S')}</p>

<div class="cards">
  <div class="card">
    <div class="card-label">Total Records</div>
    <div class="card-value blue">{stats['total']}</div>
  </div>
  <div class="card">
    <div class="card-label">Successful</div>
    <div class="card-value green">{stats['success']}</div>
  </div>
  <div class="card">
    <div class="card-label">Failed</div>
    <div class="card-value red">{stats['failed']}</div>
  </div>
  <div class="card">
    <div class="card-label">Success Rate</div>
    <div class="card-value yellow">{stats['rate']}%</div>
  </div>
</div>

<div class="table-wrap">
<table>
  <thead><tr>{headers}</tr></thead>
  <tbody>{rows_html}</tbody>
</table>
</div>

<p class="footer">RPA Bot Engine v1.0 · Python + Selenium + Pandas</p>
</body>
</html>"""

        path.write_text(html, encoding="utf-8")
        logger.info(f"🌐 HTML report saved: {path}")
        return str(path)

    # ─── Generate Both ────────────────────────────────────────────────────────

    def generate_all(self) -> dict:
        return {
            "html":  self.to_html(),
            "excel": self.to_excel(),
            "stats": self._stats(),
        }